"""Чтение сменных отчётов: листы 1–31, проверка строк 21–32 (день) и 136–147 (ночь)."""
import logging
import re
import warnings
from pathlib import Path
from typing import Iterator, List, NamedTuple

# openpyxl не поддерживает Data Validation — для чтения ячеек это не нужно
warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported",
    module="openpyxl",
)

log = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
except ImportError:
    load_workbook = None
    Workbook = None

try:
    import xlwings as xw
except ImportError:
    xw = None

# Листы по дням месяца
DAY_SHEET_NAMES = [str(d) for d in range(1, 32)]
# Строки для проверки: дневная смена 21–32, ночная 136–147. Проверяется только колонка A.
ROWS_DAY = (21, 33)
ROWS_NIGHT = (136, 148)
COLUMNS_RANGE = (1, 11)  # A–K (K = факт)
# Простои: день 47–113, ночь 162–228. A=продукт, F–G=описание, H=вид, I=начало, J=конец, K=тип (Плановый/Неплановый).
DOWNTIME_DAY = (47, 114)
DOWNTIME_NIGHT = (162, 229)
DOWNTIME_COLS = 11  # A–K


class RowDetail(NamedTuple):
    row: int
    value: str
    start: str | None
    end: str | None
    speed: str | None
    fact: str = ""  # столбец K — фактический выпуск


class DowntimeEntry(NamedTuple):
    product: str
    description: str
    kind: str
    start: str
    end: str
    type: str = ""  # Плановый / Неплановый (колонка K)


class DayInfo(NamedTuple):
    day: int
    has_data: bool
    details: List[RowDetail]
    downtimes: List[DowntimeEntry]


def _is_zero_value(val) -> bool:
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return val == 0
    if isinstance(val, str):
        normalized = val.strip().replace(",", ".")
        return normalized in {"0", "0.0", "0.00"}
    return False


def _collect_column_a(ws) -> Iterator[RowDetail]:
    """Итерирует непустые значения в колонке A для заданных строк."""
    for min_row, max_row in (ROWS_DAY, ROWS_NIGHT):
        for row_index, row in enumerate(
            ws.iter_rows(
                min_row=min_row,
                max_row=max_row - 1,
                min_col=COLUMNS_RANGE[0],
                max_col=COLUMNS_RANGE[1],
                values_only=True,
            ),
            start=min_row,
        ):
            val = row[0]
            if val is None or _is_zero_value(val):
                continue
            text = str(val).strip()
            if not text:
                continue
            start = row[1] if len(row) > 1 else None
            end = row[2] if len(row) > 2 else None
            speed = row[4] if len(row) > 4 else None
            fact_cell = row[10] if len(row) > 10 else None
            fact_str = str(fact_cell).strip() if fact_cell is not None else ""
            yield RowDetail(
                row=row_index,
                value=text,
                start=str(start).strip() if start is not None else None,
                end=str(end).strip() if end is not None else None,
                speed=str(speed).strip() if speed is not None else None,
                fact=fact_str,
            )


def _str_cell(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _collect_downtimes(ws) -> Iterator[DowntimeEntry]:
    """Простои: строки 47–113 (день), 162–228 (ночь). A, F–G, H, I, J, K (тип)."""
    for min_row, max_row in (DOWNTIME_DAY, DOWNTIME_NIGHT):
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row - 1,
            min_col=1,
            max_col=DOWNTIME_COLS,
            values_only=True,
        ):
            product = _str_cell(row[0]) if len(row) > 0 else ""
            desc_f = _str_cell(row[5]) if len(row) > 5 else ""
            desc_g = _str_cell(row[6]) if len(row) > 6 else ""
            description = desc_f or desc_g
            kind = _str_cell(row[7]) if len(row) > 7 else ""
            start = _str_cell(row[8]) if len(row) > 8 else ""
            end = _str_cell(row[9]) if len(row) > 9 else ""
            dtype = _str_cell(row[10]) if len(row) > 10 else ""
            if not description and not kind and not start and not end:
                continue
            yield DowntimeEntry(
                product=product or "—",
                description=description,
                kind=kind,
                start=start,
                end=end,
                type=dtype,
            )


def _sheet_summary(ws, day_num: int) -> DayInfo:
    items = list(_collect_column_a(ws))
    downtimes = list(_collect_downtimes(ws))
    return DayInfo(day=day_num, has_data=bool(items), details=items, downtimes=downtimes)


# Формат имени: "Сменный отчет линия № 5 - Сиропы 0,25 л. - 1,0 л.xlsx" или .xlsm
EXCEL_EXTENSIONS = (".xlsx", ".xlsm")
LINE_FILE_PREFIX = "Сменный отчет линия № "


def _line_number_from_filename(name: str) -> int | None:
    """Из имени 'Сменный отчет линия № 5 - ...' извлекает 5."""
    if not name.startswith(LINE_FILE_PREFIX):
        return None
    m = re.search(r"линия\s*№\s*(\d+)\s*\-", name, re.IGNORECASE)
    return int(m.group(1)) if m else None


def find_line_files(folder: str) -> list[tuple[int, Path]]:
    """Файлы формата 'Сменный отчет линия № N - ...' (.xlsx/.xlsm) в папке. Возвращает [(N, path), ...]."""
    if not load_workbook:
        log.warning("openpyxl не установлен")
        return []
    folder_path = Path(folder)
    if not folder_path.is_dir():
        log.error("Папка не найдена: %s", folder)
        return []
    out = []
    for path in folder_path.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() not in EXCEL_EXTENSIONS:
            continue
        n = _line_number_from_filename(path.name)
        if n is not None:
            out.append((n, path))
    out.sort(key=lambda x: x[0])
    log.info("Найдено файлов линий: %s в %s", len(out), folder)
    if not out:
        excel_in_folder = [p.name for p in folder_path.iterdir() if p.is_file() and p.suffix.lower() in EXCEL_EXTENSIONS]
        if excel_in_folder:
            log.info("В папке есть .xlsx/.xlsm (%s шт.): %s", len(excel_in_folder), ", ".join(excel_in_folder[:10]))
            if len(excel_in_folder) > 10:
                log.info("  ... и ещё %s", len(excel_in_folder) - 10)
        else:
            log.info("В папке нет .xlsx/.xlsm файлов.")
    for line_num, p in out:
        log.debug("  линия № %s: %s", line_num, p.name)
    return out


def read_line_non_empty_days(path: Path, day: int | None = None) -> list[DayInfo]:
    """По файлу линии возвращает информацию по дням. Если day задан (1–31), читается только этот лист."""
    if not load_workbook:
        return []
    log.info("Читаю файл: %s", path.name)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        log.exception("Ошибка открытия %s: %s", path.name, e)
        return []
    day_infos: List[DayInfo] = []
    sheets_to_read = [str(day)] if day is not None and 1 <= day <= 31 else DAY_SHEET_NAMES
    try:
        for name in wb.sheetnames:
            if name not in sheets_to_read:
                continue
            day_num = int(name)
            ws = wb[name]
            info = _sheet_summary(ws, day_num)
            if info.details:
                log.debug("  лист %s: найдено %s значений", name, len(info.details))
            day_infos.append(info)
    finally:
        wb.close()
    day_infos.sort(key=lambda info: info.day)
    log.info("  непустые дни: %s", [info.day for info in day_infos if info.has_data])
    return day_infos


def build_consolidated_day_excel(
    folder: str, day: int, lines_result: list, out_path: Path
) -> bool:
    """Формирует сводный Excel за день через xlwings: листы «Линия 1», «Линия 2», … с полным форматированием."""
    if not xw:
        log.warning("xlwings не установлен, сводный Excel не создаётся")
        return False
    folder_path = Path(folder).resolve()
    out_path = out_path.resolve()
    to_copy = []
    for line_num, path_name, day_infos in lines_result:
        for info in day_infos:
            if info.day == day and info.has_data:
                to_copy.append((line_num, path_name))
                break
    if not to_copy:
        log.info("Нет непустых данных за день %s, сводный Excel не создаётся", day)
        return False
    sheet_day = str(day)
    app = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.screen_updating = False
        app.display_alerts = False
        try:
            app.api.Calculation = -4135  # xlCalculationManual
        except Exception:
            pass
            
        dest = app.books.add()
        after_sheet = dest.sheets[0]
        for line_num, path_name in to_copy:
            src_path = folder_path / path_name
            if not src_path.is_file():
                log.warning("Файл не найден: %s", src_path)
                continue
            try:
                src_book = app.books.open(str(src_path))
            except Exception as e:
                log.exception("Ошибка открытия %s: %s", path_name, e)
                continue
            try:
                if sheet_day not in [s.name for s in src_book.sheets]:
                    log.warning("В файле %s нет листа «%s»", path_name, sheet_day)
                    continue
                src_sheet = src_book.sheets[sheet_day]
                n_before = len(dest.sheets)
                src_sheet.api.Copy(After=after_sheet.api)
                copied = dest.sheets[n_before]
                copied.name = f"Линия {line_num}"
                after_sheet = copied
            finally:
                src_book.close()
        if len(dest.sheets) == 1:
            log.warning("Не удалось скопировать ни одного листа")
            return False
        dest.sheets[0].delete()
        dest.save(str(out_path))
        log.info("Сводный Excel сохранён: %s", out_path)
        return True
    except Exception as e:
        log.exception("Ошибка сводного Excel %s: %s", out_path, e)
        return False
    finally:
        if app is not None:
            try:
                for b in list(app.books):
                    b.close()
            except Exception:
                pass
            try:
                app.quit()
            except Exception:
                pass
